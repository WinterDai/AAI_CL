################################################################################
# Script Name: checklist_fillin.py
#
# Purpose:
#   Fill Origin.xlsx template with module/item status and annotate waived entries.
#
# Usage:
#   from checklist_fillin import annotate_excel_template_multi, annotate_excel_template_auto
#
# Author: yyin
# Date:   2025-10-23
################################################################################
"""Back-annotate execution results into checklist artifacts.

Capabilities:
1. Annotate a plain checklist CSV (adds Checked, Comments columns).
2. Create an annotated Excel copy (Origin.xlsx) from a read-only template without modifying the original.

Annotation Rules:
	- Match items by ID (e.g. IMP-5-0-0-10) from summary YAML check_items.
	- Executed & no failures -> Checked = 'yes', Comments = ''
	- Else -> Checked = 'fail'
				* Not executed -> Comments = 'no executed'
				* Has failures -> Comments = aggregated one-line string: index detail (reason); ...
	- Items absent in summary -> leave new columns blank.

Reusable API (import from other scripts):
	annotate_checklist_csv(summary_yaml, checklist_csv, out_csv)
	annotate_excel_template(summary_yaml, template_xlsx, sheet_name, out_xlsx)

CLI retained for manual use.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
import argparse
import csv
import yaml
import sys
from datetime import datetime


def load_summary(summary_yaml: Path) -> Dict[str, Any]:
	if not summary_yaml.is_file():
		raise FileNotFoundError(f"Summary YAML not found: {summary_yaml}")
	with summary_yaml.open('r', encoding='utf-8') as f:
		return yaml.safe_load(f) or {}


def build_item_map(summary: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
	return summary.get('check_items', {}) or {}


def load_checklist(csv_path: Path) -> Tuple[List[str], List[List[str]]]:
	if not csv_path.is_file():
		raise FileNotFoundError(f"Checklist CSV not found: {csv_path}")
	with csv_path.open('r', encoding='utf-8', newline='') as f:
		r = csv.reader(f)
		rows = list(r)
	if not rows:
		return [], []
	header = rows[0]
	body = rows[1:]
	return header, body


def aggregate_failures(failures: List[Dict[str, Any]]) -> str:
	"""Aggregate failure entries into a one-line string (exclude occurrence-only dict)."""
	if not failures:
		return ''
	parts = []
	for f in failures:
		if isinstance(f, dict) and len(f) == 1 and 'occurrence' in f:
			continue  # skip occurrence record
		idx = f.get('index')
		detail = f.get('detail') or f.get('cell') or ''
		reason = f.get('reason') or ''
		seg = []
		if idx is not None:
			seg.append(str(idx))
		if detail:
			seg.append(str(detail))
		if reason:
			seg.append(f"({reason})")
		parts.append(':'.join(seg) if len(seg) <= 2 else ' '.join(seg))
	# join with ; ensuring final length not absurdly long (optional future: truncate)
	return '; '.join(parts)


def aggregate_warnings(warnings: List[Dict[str, Any]]) -> str:
	"""Aggregate warning entries into a one-line string (exclude occurrence-only dict)."""
	if not warnings:
		return ''
	parts = []
	for w in warnings:
		if isinstance(w, dict) and len(w) == 1 and 'occurrence' in w:
			continue  # skip occurrence record
		idx = w.get('index')
		detail = w.get('detail') or ''
		reason = w.get('reason') or ''
		seg = []
		if idx is not None:
			seg.append(str(idx))
		if detail:
			seg.append(str(detail))
		if reason:
			seg.append(f"({reason})")
		parts.append(':'.join(seg) if len(seg) <= 2 else ' '.join(seg))
	return '; '.join(parts)


def annotate_rows(header: List[str], rows: List[List[str]], items: Dict[str, Any]) -> Tuple[List[str], List[List[str]]]:
	# Normalize header / find indices
	col_map = {name: i for i, name in enumerate(header)}
	# We expect 'Check_modules','Item','Info'
	# Add new columns if not present
	if 'Checked' not in col_map:
		header.append('Checked')
		col_map['Checked'] = len(header) - 1
		for r in rows:
			r.append('')
	if 'Comments' not in col_map:
		header.append('Comments')
		col_map['Comments'] = len(header) - 1
		for r in rows:
			r.append('')

	item_idx = col_map.get('Item')
	if item_idx is None:
		raise ValueError("Checklist CSV missing 'Item' column")
	module_idx = col_map.get('Check_modules')

	for r in rows:
		if item_idx >= len(r):
			continue
		item_id = str(r[item_idx]).strip()
		if not item_id:
			continue
		entry = items.get(item_id)
		if entry is None:
			continue  # not in summary
		executed = str(entry.get('executed', '')).upper() in {"YES", "TRUE", "1"}
		failures = entry.get('failures') or []
		warnings = entry.get('warnings') or []
		
		# NEW LOGIC: Checked based on executed only
		checked_val = 'YES' if executed else 'NO'
		
		# Comments: aggregate failures and warnings
		comments_parts = []
		fail_str = aggregate_failures(failures)
		warn_str = aggregate_warnings(warnings)
		if fail_str:
			comments_parts.append(f"FAIL: {fail_str}")
		if warn_str:
			comments_parts.append(f"WARN: {warn_str}")
		if not executed:
			comments_parts.insert(0, "No executed")
		
		# If executed with no errors, display PASS
		if executed and not fail_str and not warn_str:
			comments_val = 'PASS'
		else:
			comments_val = ' | '.join(comments_parts) if comments_parts else ''
		
		r[col_map['Checked']] = checked_val
		r[col_map['Comments']] = comments_val
	return header, rows


def write_checklist(csv_path: Path, header: List[str], rows: List[List[str]]) -> Path:
	with csv_path.open('w', encoding='utf-8', newline='') as f:
		w = csv.writer(f)
		w.writerow(header)
		w.writerows(rows)
	return csv_path


def write_xlsx(xlsx_path: Path, header: List[str], rows: List[List[str]]) -> Path:
	"""Write an XLSX version (Origin.xlsx) preserving order and adding styles."""
	try:
		from openpyxl import Workbook  # type: ignore
		from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # type: ignore
	except Exception as e:  # noqa: BLE001
		raise RuntimeError(f"openpyxl required for XLSX output: {e}")
	wb = Workbook()
	ws = wb.active
	ws.title = 'Checklist'
	bold = Font(bold=True)
	center = Alignment(horizontal='center', vertical='center', wrap_text=True)
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	ws.append(header)
	for c in ws[1]:
		c.font = bold
		c.alignment = center
		c.border = thin_border
	checked_idx = header.index('Checked') if 'Checked' in header else -1
	comments_idx = header.index('Comments') if 'Comments' in header else -1
	# NEW COLOR SCHEME:
	# Not executed (Checked=NO) -> light green
	# Executed without errors (Checked=YES, no FAIL/WARN in Comments) -> light yellow
	# Executed with errors (Checked=YES, has FAIL or WARN) -> light red
	no_exec_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # light green
	exec_ok_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')  # light yellow
	exec_err_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # light red
	
	for r in rows:
		ws.append(r)
		last_row = ws.max_row
		if checked_idx >= 0 and checked_idx < len(r):
			checked_val = (r[checked_idx] or '').strip().upper()
			comments_val = (r[comments_idx] or '').strip() if comments_idx >= 0 and comments_idx < len(r) else ''
			
			if checked_val == 'NO':
				fill_color = no_exec_fill
			elif checked_val == 'YES':
				# Check if there are errors in comments
				if 'FAIL:' in comments_val or 'WARN:' in comments_val:
					fill_color = exec_err_fill
				else:
					fill_color = exec_ok_fill
			else:
				fill_color = None
			
			if fill_color:
				for cell in ws[last_row]:
					cell.fill = fill_color
					cell.border = thin_border
			else:
				# Still apply border even without fill color
				for cell in ws[last_row]:
					cell.border = thin_border
	# auto width
	try:
		from openpyxl.utils import get_column_letter  # type: ignore
		for col in ws.columns:
			max_len = 0
			col_letter = get_column_letter(col[0].column)
			for cell in col:
				if cell.value is None:
					continue
				l = len(str(cell.value))
				if l > max_len:
					max_len = l
			ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
	except Exception:
		pass
	meta = wb.create_sheet('Meta')
	meta.append(['Generated', datetime.now().isoformat(timespec='seconds')])
	wb.save(xlsx_path)
	return xlsx_path


def parse_args() -> argparse.Namespace:
	p = argparse.ArgumentParser(description='Back-annotate execution results into checklist CSV / XLSX.')
	p.add_argument('-summary', '--summary', dest='summary', required=True, help='Path to module summary YAML.')
	p.add_argument('-checklist', help='Path to existing CheckList.csv (for CSV annotation).')
	p.add_argument('-out-csv', help='Output annotated CSV path (if omitted and --checklist given: overwrite).')
	p.add_argument('-xlsx-template', help='Original Excel template to read (will NOT be modified).')
	p.add_argument('-xlsx-sheet', default='BE_check', help='Template sheet name (default BE_check).')
	p.add_argument('-out-xlsx', help='Output annotated Excel file path.')
	p.add_argument('--all', action='store_true', help='Aggregate all module summaries automatically (ignores single --summary item id).')
	return p.parse_args()


def _annotate_excel(summary_yamls: List[Path], template_xlsx: Path, sheet: str, out_xlsx: Path) -> Path:
	"""Core Excel annotation logic supporting single or multiple summaries.

	summary_yamls: list of YAML paths; later entries override earlier items.
	Template assumptions: header row contains 'Missing Info','Stages','Checked'.
	Column mapping: Item ID at col 5; Checked col 10; Comments col 15.
	
	IMPORTANT: Preserves ALL template formatting (borders, merges, column widths, fonts, etc.)
	Only modifies: Checked/Comments cell values and fills ONLY columns 5, 10, 15 with colors.
	"""
	from openpyxl import load_workbook  # type: ignore
	from openpyxl.styles import PatternFill  # type: ignore
	wb = load_workbook(template_xlsx)
	if sheet not in wb.sheetnames:
		raise ValueError(f"Sheet '{sheet}' not found in {template_xlsx}")
	ws = wb[sheet]
	header_row_idx: Optional[int] = None
	for row in ws.iter_rows(min_row=1, max_row=min(120, ws.max_row)):
		values = [str(c.value).strip() if c.value is not None else '' for c in row]
		if 'Missing Info' in values and 'Stages' in values and 'Checked' in values:
			header_row_idx = row[0].row
			break
	if header_row_idx is None:
		raise ValueError('Template header not found (Missing Info / Stages / Checked).')
	item_col = 5
	checked_col = 10
	auto_value_col = 12
	auto_ref_info_col = 14
	# NEW COLOR SCHEME:
	# Not executed -> light green
	# Executed without errors -> light yellow
	# Executed with errors -> light red
	no_exec_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # light green
	exec_ok_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')  # light yellow
	exec_err_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # light red
	
	combined: Dict[str, Dict[str, Any]] = {}
	for y in summary_yamls:
		summary = load_summary(y)
		combined.update(build_item_map(summary))
	for row in ws.iter_rows(min_row=header_row_idx+1):
		item_cell = row[item_col-1]
		if not item_cell or not item_cell.value:
			continue
		item_id = str(item_cell.value).strip()
		entry = combined.get(item_id)
		if not entry:
			continue
		executed = str(entry.get('executed', '')).upper() in {"YES","TRUE","1"}
		failures = entry.get('failures') or []
		warnings = entry.get('warnings') or []
		
		# NEW LOGIC: Checked based on executed only
		checked_text = 'YES' if executed else 'NO'
		
		# Auto Ref info (col 14): aggregate failures and warnings
		comments_parts = []
		fail_str = aggregate_failures(failures)
		warn_str = aggregate_warnings(warnings)
		if fail_str:
			comments_parts.append(f"FAIL: {fail_str}")
		if warn_str:
			comments_parts.append(f"WARN: {warn_str}")
		if not executed:
			comments_parts.insert(0, "No executed")
		
		# If executed with no errors, display PASS
		if executed and not fail_str and not warn_str:
			auto_ref_info_text = 'PASS'
		else:
			auto_ref_info_text = ' | '.join(comments_parts) if comments_parts else ''
		
		# Auto Value (col 12): PASS if executed without errors/warnings, FAIL otherwise
		if executed and not fail_str and not warn_str:
			auto_value_text = 'PASS'
		else:
			auto_value_text = 'FAIL'
		
		ws.cell(row=item_cell.row, column=checked_col, value=checked_text)
		ws.cell(row=item_cell.row, column=auto_value_col, value=auto_value_text)
		ws.cell(row=item_cell.row, column=auto_ref_info_col, value=auto_ref_info_text)
		
		# Determine fill color
		if not executed:
			fill_color = no_exec_fill
		elif fail_str or warn_str:
			fill_color = exec_err_fill
		else:
			fill_color = exec_ok_fill
		
		# Apply fill color to columns 5 (Item), 6, 10 (Checked), 12 (Auto Value), and 14 (Auto Ref info)
		# This preserves ALL other formatting: borders, merges, fonts, alignment, etc.
		ws.cell(row=item_cell.row, column=item_col).fill = fill_color
		ws.cell(row=item_cell.row, column=6).fill = fill_color  # Column 6
		ws.cell(row=item_cell.row, column=checked_col).fill = fill_color
		ws.cell(row=item_cell.row, column=auto_value_col).fill = fill_color
		ws.cell(row=item_cell.row, column=auto_ref_info_col).fill = fill_color
	
	# Force Excel to recalculate formulas when opening the file
	# wb.calculation.calcMode = 'auto'
	wb.save(out_xlsx)
	return out_xlsx


def annotate_excel_template(summary_yaml: Path, template_xlsx: Path, sheet: str, out_xlsx: Path) -> Path:
	return _annotate_excel([summary_yaml], template_xlsx, sheet, out_xlsx)


def annotate_excel_template_multi(summary_yamls: List[Path], template_xlsx: Path, sheet: str, out_xlsx: Path) -> Path:
	return _annotate_excel(summary_yamls, template_xlsx, sheet, out_xlsx)


def annotate_excel_template_auto(root: Path, template_xlsx: Path, sheet: str, out_xlsx: Path) -> Path:
	"""Auto-discover all module summary YAMLs under root and annotate template.

	Adds a Meta sheet listing included modules.
	"""
	summary_list: List[Path] = []
	ck_dir = root / 'Check_modules'
	if ck_dir.is_dir():
		for mod_dir in ck_dir.iterdir():
			if not mod_dir.is_dir():
				continue
			yml = mod_dir / 'outputs' / f'{mod_dir.name}.yaml'
			if yml.is_file():
				summary_list.append(yml)
	if not summary_list:
		raise ValueError('No summary YAMLs found for auto annotation.')
	result = _annotate_excel(summary_list, template_xlsx, sheet, out_xlsx)
	# Append Meta sheet with modules list
	from openpyxl import load_workbook  # type: ignore
	wb = load_workbook(result)
	if 'Meta' in wb.sheetnames:
		meta = wb['Meta']
	else:
		meta = wb.create_sheet('Meta')
	meta.append(['Included Modules'])
	for p in summary_list:
		meta.append([p.stem])
	wb.save(result)
	return result


def annotate_checklist_csv(summary_yaml: Path, checklist_csv: Path, out_csv: Path) -> Path:
	summary = load_summary(summary_yaml)
	items = build_item_map(summary)
	header, rows = load_checklist(checklist_csv)
	if not header:
		raise ValueError('Checklist CSV empty.')
	header2, rows2 = annotate_rows(header, rows, items)
	write_checklist(out_csv, header2, rows2)
	return out_csv


def main() -> int:
	args = parse_args()
	try:
		summary_path = Path(args.summary).resolve()
		# Auto root detection: ascend until Project_config exists
		def detect_root(start: Path) -> Path:
			cur = start
			for _ in range(10):  # limit ascent
				if (cur / 'Project_config').is_dir():
					return cur
				if cur.parent == cur:
					break
				cur = cur.parent
			return start.parent  # fallback
		project_root = detect_root(summary_path.parent)

		# If user didn't supply template/output, auto-fill
		if not args.xlsx_template:
			auto_template = project_root / 'Project_config' / 'collaterals' / 'Initial' / 'latest' / 'DR3_SSCET_BE_Check_List_v0.1.xlsx'
		else:
			auto_template = Path(args.xlsx_template).resolve()
		if not args.out_xlsx:
			auto_out_xlsx = project_root / 'Work' / 'Results' / 'Origin.xlsx'
		else:
			auto_out_xlsx = Path(args.out_xlsx).resolve()
		sheet_name = args.xlsx_sheet

		# Build list of summary yamls if --all
		if args.all:
			summaries: List[Path] = []
			check_modules_dir = project_root / 'Check_modules'
			if check_modules_dir.is_dir():
				for mod_dir in check_modules_dir.iterdir():
					if not mod_dir.is_dir():
						continue
					out_yaml = mod_dir / 'outputs' / f'{mod_dir.name}.yaml'
					if out_yaml.is_file():
						summaries.append(out_yaml.resolve())
			if not summaries:
				raise ValueError('No module summary YAMLs found for aggregation.')
			annotate_excel_template_multi(summaries, auto_template, sheet_name, auto_out_xlsx)
			print(f"[INFO] Aggregated XLSX written: {auto_out_xlsx}")
		else:
			# Single summary annotation
			if auto_template.is_file():
				annotate_excel_template(summary_path, auto_template, sheet_name, auto_out_xlsx)
				print(f"[INFO] XLSX written: {auto_out_xlsx}")
			else:
				raise FileNotFoundError(f'Template Excel not found: {auto_template}')
		if args.checklist:
			checklist_path = Path(args.checklist).resolve()
			out_csv = Path(args.out_csv).resolve() if args.out_csv else checklist_path
			annotate_checklist_csv(summary_path, checklist_path, out_csv)
			print(f"[INFO] CSV annotated: {out_csv}")
		return 0
	except Exception as e:
		print(f"[ERROR] {e}")
		return 2


if __name__ == '__main__':  # pragma: no cover
	sys.exit(main())

