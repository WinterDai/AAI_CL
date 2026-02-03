"""Utilities to parse a master checklist Excel file into per-sheet Excel files
and a consolidated CSV extracted from the 'BE_check' sheet.
"""

from __future__ import absolute_import, division, print_function

from pathlib import Path
import re
import sys
from typing import Optional, List, Dict, Any

# Attempt to import pandas; in Python 3.6 an incompatible newer pandas (>=2.0) may raise SyntaxError
pd = None  # type: ignore
try:  # noqa: E722 (broad for SyntaxError/ImportError)
    import pandas as _pd  # type: ignore
    # Guard: pandas 2.x requires Python >=3.8, so if running 3.6 and version looks 2.*, ignore
    ver = getattr(_pd, '__version__', '0')
    if sys.version_info < (3, 7) and ver.startswith('2.'):
        _pd = None  # force fallback
    pd = _pd
except Exception:  # includes SyntaxError when encountering unsupported future imports
    pd = None

try:
    import openpyxl  # Fallback reader/writer
except ImportError:
    openpyxl = None  # type: ignore


def parse_in(stage, excel_file, output_dir=None):  # type: (str, str, Optional[str]) -> str
    """Parse an input Excel workbook.

    Steps:
      1. Save every sheet as its own .xlsx in output_dir
      2. From sheet named 'BE_check' (case insensitive), extract checklist items:
         - A module row is identified by column 4 matching leading number pattern.
         - Following rows with non-empty columns 4/5 become checklist entries until next module.
      3. Emit consolidated CSV 'CheckList.csv'.

    Args:
        stage (str): Flow stage identifier (currently unused, kept for interface stability).
        excel_file (str): Path to source Excel file.
        output_dir (str|None): Directory to place outputs (defaults to excel file parent).

    Returns:
        str: Path to generated CSV file.
    """
    excel_path = Path(excel_file)
    if output_dir is None:
        output_dir = str(excel_path.parent)
    out_dir_path = Path(output_dir)
    if not out_dir_path.exists():
        # Python 3.6 safe directory creation
        out_dir_path.mkdir(parents=True)

    # Explicit engine for deterministic behavior; openpyxl supports .xlsx
    # Branch 1: pandas path
    if pd is not None:
        try:
            xls = pd.ExcelFile(excel_path, engine='openpyxl')
        except Exception:
            xls = None
        if xls is not None:
            # Save each sheet
            for sheet_name in xls.sheet_names:
                try:
                    df_sheet = xls.parse(sheet_name)
                except Exception:
                    continue
                safe_sheet_filename = "{}.xlsx".format(sheet_name)
                try:
                    df_sheet.to_excel(out_dir_path / safe_sheet_filename, index=False)
                except Exception:
                    pass

            # Find target sheet case-insensitively
            be_sheet_name = None
            for name in xls.sheet_names:
                if name.lower() == 'be_check':
                    be_sheet_name = name
                    break

            if be_sheet_name is None:
                empty_csv_path = out_dir_path / 'CheckList.csv'
                pd.DataFrame([], columns=['Check_modules', 'Item', 'Info']).to_csv(empty_csv_path, index=False)
                return str(empty_csv_path)

            df_be_check = xls.parse(be_sheet_name, header=None)

            checklist = []  # type: List[Dict[str, Any]]
            current_module = None  # type: Optional[str]
            module_re = re.compile(r"^\s*\d+\.?\d*\b")

            def normalize_module_name(s):  # type: (str) -> str
                out = re.sub(r"\([^)]*\)", "", s)
                out = re.sub(r"[-/\s]+", "_", out)
                out = re.sub(r"_+", "_", out)
                return out.strip("_")

            for _, row in df_be_check.iterrows():
                module_cell = None
                if len(row) > 4 and pd.notna(row[4]):
                    try:
                        module_cell = str(row[4]).strip()
                    except Exception:
                        module_cell = None

                if module_cell and module_re.search(module_cell):
                    current_module = normalize_module_name(module_cell)
                    continue

                if current_module is not None:
                    if len(row) <= 4 or (pd.isna(row[4]) and (len(row) <= 5 or pd.isna(row[5]))):
                        continue

                    item = ''
                    info = ''
                    if len(row) > 4 and pd.notna(row[4]):
                        item = str(row[4]).strip()
                    if len(row) > 5 and pd.notna(row[5]):
                        info = str(row[5]).strip()

                    if item and info and item.lower() != 'nan' and info.lower() != 'nan':
                        checklist.append({
                            'Check_modules': current_module,
                            'Item': item,
                            'Info': info
                        })

            out_csv_path = out_dir_path / 'CheckList.csv'
            pd.DataFrame(checklist).to_csv(out_csv_path, index=False)
            return str(out_csv_path)

    # Branch 2: openpyxl fallback (no pandas)
    if openpyxl is None:
        raise RuntimeError("Neither pandas nor openpyxl is available; cannot parse Excel.")

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    # Save each sheet as its own workbook
    for sheet_name in wb.sheetnames:
        try:
            ws_src = wb[sheet_name]
            new_wb = openpyxl.Workbook()
            ws_new = new_wb.active
            ws_new.title = sheet_name
            for row in ws_src.iter_rows(values_only=True):
                ws_new.append(list(row))
            safe_sheet_filename = "{}.xlsx".format(sheet_name)
            new_wb.save(str(out_dir_path / safe_sheet_filename))
        except Exception:
            continue

    be_sheet_name = None
    for name in wb.sheetnames:
        if name.lower() == 'be_check':
            be_sheet_name = name
            break

    import csv
    out_csv_path = out_dir_path / 'CheckList.csv'

    if be_sheet_name is None:
        # Create empty CSV header
        with open(str(out_csv_path), 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Check_modules', 'Item', 'Info'])
        return str(out_csv_path)

    ws = wb[be_sheet_name]

    module_re = re.compile(r"^\s*\d+\.?\d*\b")
    current_module = None

    def normalize_module_name(s):  # type: (str) -> str
        out = re.sub(r"\([^)]*\)", "", s)
        out = re.sub(r"[-/\s]+", "_", out)
        out = re.sub(r"_+", "_", out)
        return out.strip("_")

    rows_out = []
    for row in ws.iter_rows(values_only=True):
        row_list = list(row)
        # Access columns 5 and 6 (0-based 4,5)
        col4 = row_list[4] if len(row_list) > 4 else None
        col5 = row_list[5] if len(row_list) > 5 else None

        module_cell = None
        if col4 is not None and str(col4).strip():
            try:
                module_cell = str(col4).strip()
            except Exception:
                module_cell = None

        if module_cell and module_re.search(module_cell):
            current_module = normalize_module_name(module_cell)
            continue

        if current_module is not None:
            if (col4 is None or str(col4).strip().lower() == 'nan') and (col5 is None or str(col5).strip().lower() == 'nan'):
                continue
            item = '' if col4 is None else str(col4).strip()
            info = '' if col5 is None else str(col5).strip()
            if item and info and item.lower() != 'nan' and info.lower() != 'nan':
                rows_out.append([current_module, item, info])

    with open(str(out_csv_path), 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Check_modules', 'Item', 'Info'])
        writer.writerows(rows_out)

    return str(out_csv_path)