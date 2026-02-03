################################################################################
# Script Name: excel_summary_generator.py
#
# Purpose:
#   Build global Summary.xlsx and supporting sheets from all module summary YAMLs.
#   Aggregates Overview metrics, per-module breakdown, and waiver information.
#
# Usage:
#   python excel_summary_generator.py --out <OUT_PATH> --root <ROOT>
#   (Normally invoked by check_flowtool.py after modules executed)
#
# Author: yyin
# Date:   2025-10-23
################################################################################
r"""Generate aggregated Summary.xlsx across all module summary YAMLs.

Usage (PowerShell):
  python excel_summary_generator.py
	python excel_summary_generator.py --out C:/Users/yuyin/Desktop/CHECKLIST/Work/Results/Summary.xlsx

Discovery:
  Scans sibling module folders under Check_modules/*/outputs/*.yaml.
  Each summary YAML expected to have structure:
	module: <name>
	check_items: { item_id: { executed: bool, status: str, failures: [..]?, warnings: [..]? } }

Sheet structure:
  Overview: Aggregated metrics (Modules, Modules_Checked, Total_Items, etc.)
  Summary: Concatenated Summary data from all module xlsx files
  Check_Info: Per-module breakdown (same as old Module_Breakdown)
  Waive_Info: Parsed waive.yaml entries from each module's Waiver_config

Warn Logic:
  If an item has any warnings (non-occurrence entries), it's counted as WARN regardless of status field.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, Any, List
import yaml

try:
	from openpyxl import Workbook  # type: ignore
	from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # type: ignore
except ImportError as e:  # pragma: no cover
	raise SystemExit("openpyxl is required. Install with: pip install openpyxl") from e


def find_workspace_root(start: Path) -> Path:
	"""Ascend directories until a folder containing 'Check_modules' is found."""
	current = start.resolve()
	while current != current.parent:
		if (current / 'Check_modules').is_dir():
			return current
		current = current.parent
	return start.resolve()


def discover_summary_yamls(root: Path) -> List[Path]:
	check_modules = root / 'Check_modules'
	yamls: List[Path] = []
	if not check_modules.is_dir():
		return yamls
	for module_dir in check_modules.iterdir():
		if not module_dir.is_dir():
			continue
		outputs = module_dir / 'outputs'
		if outputs.is_dir():
			for y in outputs.glob('*.yaml'):
				yamls.append(y)
	return sorted(yamls)


def load_summary(path: Path) -> Dict[str, Any]:
	with path.open('r', encoding='utf-8') as f:
		return yaml.safe_load(f) or {}


def has_warnings(item_info: Dict[str, Any]) -> bool:
	"""Check if item has any warnings (non-empty warnings list excluding occurrence-only dict)."""
	warnings = item_info.get('warnings') or []
	for w in warnings:
		if isinstance(w, dict) and not (len(w) == 1 and 'occurrence' in w):
			return True
	return False


def compute_stats(summary: Dict[str, Any]) -> Dict[str, Any]:
	"""Compute per-module statistics.

	Warn logic: Item counts as warn if it has non-empty warnings list (regardless of status field).
	failure_entries and warning_entries exclude the occurrence-only dict element.
	"""
	items = summary.get('check_items', {}) or {}
	executed = 0
	not_executed = 0
	passed = 0
	failed = 0
	warned = 0
	failure_entries = 0
	warning_entries = 0
	for _item_id, info in items.items():
		ex = bool(info.get('executed'))
		status = str(info.get('status', '')).lower()
		fails = info.get('failures') or []
		warnings = info.get('warnings') or []
		# Count only detailed failure dicts (exclude single-key occurrence dict)
		for f in fails:
			if isinstance(f, dict) and not (len(f) == 1 and 'occurrence' in f):
				failure_entries += 1
		for w in warnings:
			if isinstance(w, dict) and not (len(w) == 1 and 'occurrence' in w):
				warning_entries += 1
		if ex:
			executed += 1
			# Count based on actual status field
			if status == 'pass':
				passed += 1
			elif status.startswith('fail'):
				failed += 1
			# ADDITIONAL: if item has warnings, also count as warn (doesn't override status)
			if has_warnings(info):
				warned += 1
		else:
			not_executed += 1
	return {
		'module': summary.get('module', 'UNKNOWN'),
		'items_total': len(items),
		'executed': executed,
		'not_executed': not_executed,
		'pass': passed,
		'fail': failed,
		'warn': warned,
		'failure_entries': failure_entries,
		'warning_entries': warning_entries,
	}


def aggregate(all_stats: List[Dict[str, Any]]) -> Dict[str, Any]:
	total_items_sum = 0
	executed = 0
	not_executed = 0
	passed = 0
	failed = 0
	warned = 0
	failure_entries = 0
	warning_entries = 0
	for s in all_stats:
		total_items_sum += s['items_total']
		executed += s['executed']
		not_executed += s['not_executed']
		passed += s['pass']
		failed += s['fail']
		warned += s['warn']
		failure_entries += s['failure_entries']
		warning_entries += s['warning_entries']
	return {
		'modules': len(all_stats),
		'items_total_sum': total_items_sum,
		'executed': executed,
		'not_executed': not_executed,
		'pass': passed,
		'fail': failed,
		'warn': warned,
		'failure_entries': failure_entries,
		'warning_entries': warning_entries,
	}


def compute_all_stats(summary_yamls: List[Path]) -> tuple[List[Dict[str, Any]], Dict[str, Any]]:
	module_stats: List[Dict[str, Any]] = []
	unique_items: set[str] = set()
	for p in summary_yamls:
		summary = load_summary(p)
		items = summary.get('check_items', {}) or {}
		unique_items.update(items.keys())
		module_stats.append(compute_stats(summary))
	agg = aggregate(module_stats)
	agg['unique_items'] = len(unique_items)
	return module_stats, agg


def auto_size(ws) -> None:
	for col in ws.columns:
		max_len = 0
		col_letter = col[0].column_letter
		for cell in col:
			val = cell.value
			if val is None:
				continue
			max_len = max(max_len, len(str(val)))
		ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def build_summary_rows(summary_yamls: List[Path]) -> List[List[Any]]:
	"""Build Summary sheet rows from all module summary YAMLs.
	
	Columns: Module, Stage, ItemID, Executed, Status, Description, Kind, Occurrence, Index, Detail, SourceLine, SourceFile, Reason
	"""
	rows = []
	for yaml_path in summary_yamls:
		summary = load_summary(yaml_path)
		module = summary.get('module', 'UNKNOWN')
		stage = summary.get('stage', '')
		items = summary.get('check_items', {}) or {}
		
		for item_id, info in items.items():
			executed = info.get('executed')
			# Keep original status (don't override)
			status = str(info.get('status', '')).upper()
			description = info.get('description', '')
			
			# Process failures
			failures = info.get('failures') or []
			fail_occ = ''
			fail_details = []
			for f in failures:
				if isinstance(f, dict):
					if 'occurrence' in f and len(f) == 1:
						fail_occ = f['occurrence']
					else:
						fail_details.append(f)
			
			# Process warnings
			warnings_list = info.get('warnings') or []
			warn_occ = ''
			warn_details = []
			for w in warnings_list:
				if isinstance(w, dict):
					if 'occurrence' in w and len(w) == 1:
						warn_occ = w['occurrence']
					else:
						warn_details.append(w)
			
			# Add failure rows
			if fail_details:
				for fail in fail_details:
					rows.append([
						module, stage, item_id, executed, status, description, 'FAILURE', fail_occ,
						fail.get('index', ''), fail.get('detail', ''), fail.get('source_line', ''),
						fail.get('source_file', ''), fail.get('reason', '')
					])
			
			# Add warning rows
			if warn_details:
				for warn in warn_details:
					rows.append([
						module, stage, item_id, executed, status, description, 'WARNING', warn_occ,
						warn.get('index', ''), warn.get('detail', ''), warn.get('source_line', ''),
						warn.get('source_file', ''), warn.get('reason', '')
					])
			
			# If no details, add single row
			if not fail_details and not warn_details:
				rows.append([module, stage, item_id, executed, status, description, '', '', '', '', '', '', ''])
	
	return rows


def parse_waive_yaml(root: Path, modules: List[str]) -> List[List[Any]]:
	"""Parse waive.yaml files from Waiver_config/<module>/waive.yaml.

	Return rows: [Module, Description, Cell, Reason]

	Parsing strategy:
	1. Try YAML load. If top-level is a mapping, each key is treated as Description (e.g. waive_dont_use_cells)
	   and its list items are cell entries possibly containing ';' for reason.
	2. Fallback to line-based parsing (legacy) if YAML load fails.
	"""
	rows: List[List[Any]] = []
	for module in modules:
		waive_path = root / 'Waiver_config' / module / 'waive.yaml'
		if not waive_path.is_file():
			continue
		text = waive_path.read_text(encoding='utf-8', errors='ignore')
		used_structured = False
		try:
			data = yaml.safe_load(text)
			if isinstance(data, dict):
				for desc, items in data.items():
					if isinstance(items, list):
						for raw in items:
							if not isinstance(raw, str):
								continue
							entry = raw.strip()
							if not entry:
								continue
							cell_part, reason_part = entry, ''
							if ';' in entry:
								cell_part, reason_part = entry.split(';', 1)
							cell_part = cell_part.strip()
							reason_part = reason_part.strip().lstrip('#').strip()
							if cell_part:
								rows.append([module, desc, cell_part, reason_part])
				used_structured = True
		except Exception:
			used_structured = False
		if used_structured:
			continue
		# Fallback legacy parsing by lines if structured failed
		for ln_no, raw in enumerate(text.splitlines(), start=1):
			s = raw.strip()
			if not s or s.endswith(':'):
				continue
			if s.startswith('-'):
				s = s[1:].strip()
			if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
				s = s[1:-1]
			cell_part, reason_part = s, ''
			if ';' in s:
				cell_part, reason_part = s.split(';', 1)
			cell_part = cell_part.strip()
			reason_part = reason_part.strip().lstrip('#').strip()
			if cell_part:
				rows.append([module, '(line)', cell_part, reason_part])
	return rows


def write_summary_excel(out_xlsx: Path, module_stats: List[Dict[str, Any]], agg: Dict[str, Any], 
                        modules_included: List[str], summary_yamls: List[Path], root: Path) -> Path:
	wb = Workbook()
	bold = Font(bold=True)
	center = Alignment(horizontal='center', vertical='center', wrap_text=True)
	fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
	pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
	warn_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)

	# Sheet 1: Overview
	ws_overview = wb.active
	ws_overview.title = 'Overview'
	
	rows = [
		('Modules', agg['modules']),
		('Modules_Checked', ', '.join(modules_included)),
		('Total_Items', agg['items_total_sum']),
		('Executed_Items', agg['executed']),
		('Not_Executed_Items', agg['not_executed']),
		('Pass', agg['pass']),
		('Fail', agg['fail']),
		('Warn', agg['warn']),
	]
	ws_overview.append(['Metric', 'Value'])
	for cell in ws_overview[1]:
		cell.font = bold
		cell.alignment = center
		cell.border = thin_border
	for r in rows:
		ws_overview.append(r)
	# Apply borders to all data rows
	for row in ws_overview.iter_rows(min_row=2, max_row=ws_overview.max_row):
		for cell in row:
			cell.border = thin_border
	auto_size(ws_overview)

	# Sheet 2: Summary (aggregated from all module Summary sheets)
	ws_summary = wb.create_sheet('Summary')
	summary_headers = ['Module', 'Stage', 'ItemID', 'Executed', 'Status', 'Description', 'Kind', 
	                   'Occurrence', 'Index', 'Detail', 'SourceLine', 'SourceFile', 'Reason']
	ws_summary.append(summary_headers)
	for cell in ws_summary[1]:
		cell.font = bold
		cell.alignment = center
		cell.border = thin_border
	
	summary_rows = build_summary_rows(summary_yamls)
	for row in summary_rows:
		ws_summary.append(row)
	
	# Apply conditional formatting to Status column
	from openpyxl.formatting.rule import CellIsRule
	from openpyxl.utils import get_column_letter
	status_col_idx = summary_headers.index('Status') + 1
	status_col_letter = get_column_letter(status_col_idx)
	status_range = f"{status_col_letter}2:{status_col_letter}{ws_summary.max_row}"
	ws_summary.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"PASS"'], fill=pass_fill))
	ws_summary.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"FAIL"'], fill=fail_fill))
	ws_summary.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"WARN"'], fill=warn_fill))
	
	# Row highlight for PASS, WARN and FAIL rows and apply borders
	for r in range(2, ws_summary.max_row + 1):
		cell_val = ws_summary.cell(row=r, column=status_col_idx).value
		if cell_val:
			cell_val_upper = str(cell_val).upper()
			if cell_val_upper.startswith("FAIL"):
				for c in ws_summary[r]:
					c.fill = fail_fill
					c.border = thin_border
			elif cell_val_upper == "PASS":
				for c in ws_summary[r]:
					c.fill = pass_fill
					c.border = thin_border
			elif cell_val_upper == "WARN":
				for c in ws_summary[r]:
					c.fill = warn_fill
					c.border = thin_border
			else:
				for c in ws_summary[r]:
					c.border = thin_border
		else:
			for c in ws_summary[r]:
				c.border = thin_border
	
	ws_summary.freeze_panes = "A2"
	ws_summary.auto_filter.ref = ws_summary.dimensions
	auto_size(ws_summary)

	# Sheet 3: Check_Info (module breakdown, same as old Module_Breakdown)
	ws_check = wb.create_sheet('Check_Info')
	headers = ['Module', 'Items', 'Executed', 'Not Executed', 'Pass', 'Fail', 'Warn', 'Failure Entries', 'Warning Entries']
	ws_check.append(headers)
	for cell in ws_check[1]:
		cell.font = bold
		cell.alignment = center
		cell.border = thin_border
	for s in module_stats:
		ws_check.append([
			s['module'], s['items_total'], s['executed'], s['not_executed'], s['pass'], s['fail'], s['warn'], 
			s['failure_entries'], s['warning_entries']
		])
	# Color rows: pass rows (all executed and no fails) green; rows with fails red; rows with warns yellow and apply borders
	for row in ws_check.iter_rows(min_row=2, max_row=ws_check.max_row):
		fail_cnt = row[5].value
		warn_cnt = row[6].value
		if fail_cnt and fail_cnt > 0:
			for c in row:
				c.fill = fail_fill
				c.border = thin_border
		elif warn_cnt and warn_cnt > 0:
			for c in row:
				c.fill = warn_fill
				c.border = thin_border
		else:
			# Check if all executed (pass rows)
			exec_val = row[2].value
			items_val = row[1].value
			if exec_val and items_val and int(exec_val) == int(items_val):
				for c in row:
					c.fill = pass_fill
					c.border = thin_border
			else:
				for c in row:
					c.border = thin_border
	auto_size(ws_check)

	# Sheet 4: Waive_Info
	ws_waive = wb.create_sheet('Waive_Info')
	waive_headers = ['Module', 'Description', 'Cell', 'Reason']
	ws_waive.append(waive_headers)
	for cell in ws_waive[1]:
		cell.font = bold
		cell.alignment = center
		cell.border = thin_border
	
	waive_rows = parse_waive_yaml(root, modules_included)
	if waive_rows:
		for row in waive_rows:
			ws_waive.append(row)
	else:
		ws_waive.append(['No waive data found', '', '', ''])
	# Apply borders to all data rows
	for row in ws_waive.iter_rows(min_row=2, max_row=ws_waive.max_row):
		for cell in row:
			cell.border = thin_border
	auto_size(ws_waive)

	out_xlsx.parent.mkdir(parents=True, exist_ok=True)
	wb.save(out_xlsx)
	return out_xlsx


def build(out: Path | None = None, root: Path | None = None, summary_yamls: List[Path] | None = None) -> Path:
	if root is None:
		root = find_workspace_root(Path.cwd())
	else:
		root = root.resolve()
	if summary_yamls is None:
		yamls = discover_summary_yamls(root)
	else:
		yamls = summary_yamls
	if not yamls:
		raise SystemExit('No summary YAMLs found to aggregate')
	module_stats, agg = compute_all_stats(yamls)
	if out is None:
		out = root / 'Work' / 'Results' / 'Summary.xlsx'
	# Derive module names for metadata
	modules_included = [p.stem for p in yamls]
	path = write_summary_excel(out, module_stats, agg, modules_included, yamls, root)
	print(f'[INFO] Aggregated Summary written: {path}')
	return path


def _cli():
	ap = argparse.ArgumentParser(description='Generate aggregated Summary.xlsx across module summaries')
	ap.add_argument('--out', type=Path, help='Output Summary.xlsx path (default Work/Results/Summary.xlsx)')
	ap.add_argument('--root', type=Path, help='Root path containing Check_modules (auto-detect if omitted)')
	args = ap.parse_args()
	build(args.out, args.root)


if __name__ == '__main__':  # pragma: no cover
	_cli()
