"""
Dashboard API for developer status and real-time activity monitoring.
"""

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional, List
import subprocess
import os
from pathlib import Path
from datetime import datetime
import json
import re
import httpx
import asyncio
import yaml
import openpyxl

router = APIRouter(prefix="/api/dashboard", tags=["dashboard"])

# Workspaces directory path - try multiple locations
# 1. Environment variable
# 2. Relative to project root (for developers)
# 3. Default admin path (for yuyin)
def get_workspaces_dir():
    """
    智能检测工作目录：
    - Admin 环境：返回 Mdispatcher/workspaces 目录（包含所有开发者 workspace）
    - Developer 环境：返回包含当前 workspace 的父目录（模拟 workspaces 结构）
    """
    # 1. 环境变量优先
    env_path = os.environ.get("WORKSPACES_DIR")
    if env_path and os.path.exists(env_path):
        return env_path
    
    # 2. 检测 Admin 环境
    # 从 Tool/AutoGenChecker/web_ui/backend/api/dashboard.py 向上到 CHECKLIST_V4/
    # parents[5] = Tool/, parents[6] = CHECKLIST_V4/
    project_root = Path(__file__).resolve().parents[5]  # 到达 CHECKLIST_V4
    admin_workspaces = project_root / "Tool" / "Mdispatcher" / "workspaces"
    if admin_workspaces.exists():
        return str(admin_workspaces)
    
    # 3. 检测 Developer 环境
    # 从 workspace/Tool/AutoGenChecker/web_ui/backend/api/dashboard.py
    # parents[5] = Tool/, parents[6] = <workspace_root>/, parents[7] = workspaces/
    workspace_root = Path(__file__).resolve().parents[6]
    
    # 验证是否是有效的 workspace（检查特征目录）
    if (workspace_root / "CHECKLIST").exists() and (workspace_root / ".git").exists():
        # Developer 环境：返回包含当前 workspace 的父目录
        # 这样 get_all_developers() 能找到当前 workspace 目录
        return str(workspace_root.parent)
    
    # 4. Fallback（向后兼容）
    default_path = r"C:\Users\yuyin\Desktop\CHECKLIST_V4\Tool\Mdispatcher\workspaces"
    return default_path

WORKSPACES_DIR = get_workspaces_dir()

# Re-assign Excel path (with item status)
# Try environment variable first, then fallback to default path
# Other users can set REASSIGN_EXCEL_PATH environment variable to their own path
def get_excel_path():
    """
    智能检测 Excel 文件路径：
    - 优先使用 OneDrive 共享路径（适用于所有有权限的用户）
    - 其次查找本地项目副本
    """
    # 1. 环境变量优先
    env_path = os.environ.get("REASSIGN_EXCEL_PATH")
    if env_path and os.path.exists(env_path):
        return env_path
    
    # 2. 使用 OneDrive 环境变量构建共享路径（适用于所有有权限的开发者）
    # Windows OneDrive 商业版环境变量
    onedrive_commercial = os.environ.get("OneDriveCommercial")
    if not onedrive_commercial:
        # 尝试通用 OneDrive 环境变量
        onedrive_commercial = os.environ.get("OneDrive")
    
    if onedrive_commercial:
        # 如果环境变量指向个人 OneDrive，尝试构建商业版路径
        if "OneDrive - Cadence Design Systems Inc" not in onedrive_commercial:
            # 尝试在同级目录查找商业版 OneDrive
            onedrive_parent = Path(onedrive_commercial).parent
            commercial_onedrive = onedrive_parent / "OneDrive - Cadence Design Systems Inc"
            if commercial_onedrive.exists():
                onedrive_commercial = str(commercial_onedrive)
        
        shared_excel = Path(onedrive_commercial) / "2025" / "Checklist" / "RELEASE" / "DOC" / "DR3_SSCET_BE_Check_List_v0.2_20260105.xlsx"
        if shared_excel.exists():
            return str(shared_excel)
    
    # 3. 检测 Developer 环境中的副本（workspace/Tool/AutoGenChecker/）
    workspace_root = Path(__file__).resolve().parents[6]
    dev_excel = workspace_root / "Tool" / "AutoGenChecker" / "DR3_SSCET_BE_Check_List_v0.2_20260105.xlsx"
    if dev_excel.exists():
        return str(dev_excel)
    
    # 4. 检测 Admin 环境中的项目副本（CHECKLIST_V4/Tool/AutoGenChecker/）
    project_root = Path(__file__).resolve().parents[5]
    admin_excel = project_root / "Tool" / "AutoGenChecker" / "DR3_SSCET_BE_Check_List_v0.2_20260105.xlsx"
    if admin_excel.exists():
        return str(admin_excel)
    
    # 5. 项目根目录
    relative_excel = project_root / "DR3_SSCET_BE_Check_List_v0.2_20260105.xlsx"
    if relative_excel.exists():
        return str(relative_excel)
    
    # 6. Fallback to hardcoded path (yuyin's OneDrive)
    default_path = r"C:\Users\yuyin\OneDrive - Cadence Design Systems Inc\2025\Checklist\RELEASE\DOC\DR3_SSCET_BE_Check_List_v0.2_20260105.xlsx"
    return default_path

REASSIGN_EXCEL_PATH = get_excel_path()

# GitHub Enterprise API configuration
GITHUB_API_BASE = "https://github.cadence.com/api/v3"
GITHUB_REPO = "yuyin/CHECKLIST"
# Try to get token from environment or git config
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")


class DeveloperStatus(BaseModel):
    """Developer status model"""
    developer_name: str
    workspace_name: str
    current_branch: str
    last_commit_time: Optional[str] = None
    last_commit_message: Optional[str] = None
    modified_files_count: int = 0
    modified_items: List[str] = []  # List of item IDs
    status: str = "idle"  # active, idle, working
    assigned_item: Optional[str] = None


class ActivityItem(BaseModel):
    """Activity item model"""
    id: str
    developer: str
    action: str
    item_id: Optional[str] = None
    module: Optional[str] = None
    status: str
    time: str
    workspace: str


class DashboardStats(BaseModel):
    """Dashboard statistics model"""
    total_developers: int
    active_developers: int
    completed_today: int
    total_commits_today: int


def parse_developer_name(workspace_name: str) -> str:
    """Extract developer name from workspace folder name"""
    # Format: developer_name_workspace_timestamp
    match = re.match(r"(.+?)_workspace_\d+_\d+", workspace_name)
    if match:
        name = match.group(1)
        # Convert snake_case to readable name
        return name.replace("_", " ").title()
    return workspace_name


def convert_excel_module_to_directory_format(excel_module: str) -> str:
    """Convert Excel module name to Check_modules directory format
    
    Example:
        ' 2.0 - TECHFILE AND RULE DECK CHECK (GENERAL FOR ALL IPS)' -> '2.0_TECHFILE_AND_RULE_DECK_CHECK'
        ' 8.1 - PHYSICAL IMPLEMENTATION CHECK (ONLY FOR MIG HSPHY/HPPHY/DFPHY/GDDRPHY)' -> '8.1_PHYSICAL_IMPLEMENTATION_CHECK'
    """
    if not excel_module:
        return "Unknown"
    
    # Strip whitespace and convert to uppercase
    module = excel_module.strip().upper()
    
    # Extract module number and name
    # Pattern: "X.Y - NAME (description)"
    match = re.match(r'^([\d\.]+)\s*-\s*([^(]+)', module)
    if match:
        module_num = match.group(1)  # e.g., "2.0" or "8.1"
        module_name = match.group(2).strip()  # e.g., "TECHFILE AND RULE DECK CHECK"
        
        # Remove "(GENERAL FOR ALL IPS)" or "(ONLY FOR MIG...)" suffix if present
        # by taking only the first part before any trailing description
        module_name = module_name.split('(')[0].strip()
        
        # Replace spaces with underscores
        module_name = module_name.replace(" ", "_")
        
        # Format: X.Y_MODULE_NAME
        return f"{module_num}_{module_name}"
    
    return "Unknown"


def get_assigned_items(workspace_path: str, developer_name_snake: str) -> List[str]:
    """Read assigned items from assignment YAML file"""
    # Look for assignment file in workspace/CHECKLIST/Work/assignments/
    assignments_dir = os.path.join(workspace_path, "CHECKLIST", "Work", "assignments")
    if not os.path.exists(assignments_dir):
        return []
    
    # Find assignment file for this developer
    try:
        for filename in os.listdir(assignments_dir):
            if filename.startswith(developer_name_snake) and filename.endswith(".yaml"):
                assignment_file = os.path.join(assignments_dir, filename)
                with open(assignment_file, 'r', encoding='utf-8') as f:
                    data = yaml.safe_load(f)
                    if data and 'all_items' in data:
                        return sorted(data['all_items'])
    except Exception as e:
        print(f"Error reading assignment file for {developer_name_snake}: {e}")
    
    return []


# Cache for item status from Re-assign Excel
_item_status_cache = {}
_item_status_cache_time = None
_item_status_file_mtime = None  # Track file modification time
ITEM_STATUS_CACHE_TTL = 60  # Cache for 1 minute (reduced for debugging)


def clear_item_status_cache():
    """Clear the item status cache to force reload"""
    global _item_status_cache, _item_status_cache_time, _item_status_file_mtime
    _item_status_cache = {}
    _item_status_cache_time = None
    _item_status_file_mtime = None


def get_item_status_from_excel() -> dict:
    """Read item status from Re-assign Excel sheet"""
    global _item_status_cache, _item_status_cache_time, _item_status_file_mtime
    
    # Check if file exists and get modification time
    if not os.path.exists(REASSIGN_EXCEL_PATH):
        print(f"Re-assign Excel not found: {REASSIGN_EXCEL_PATH}")
        return {}
    
    current_mtime = os.path.getmtime(REASSIGN_EXCEL_PATH)
    
    # Check cache - invalidate if file was modified or cache expired
    now = datetime.now()
    cache_valid = (
        _item_status_cache_time and 
        (now - _item_status_cache_time).total_seconds() < ITEM_STATUS_CACHE_TTL and
        _item_status_file_mtime == current_mtime  # File not changed
    )
    
    if cache_valid:
        return _item_status_cache
    
    item_status = {}
    temp_file = None
    
    try:
        # Copy to temp file to avoid lock issues when Excel has the file open
        import shutil
        import tempfile
        temp_dir = tempfile.gettempdir()
        import uuid
        temp_file = os.path.join(temp_dir, f"reassign_excel_{uuid.uuid4().hex}.xlsx")
        shutil.copy2(REASSIGN_EXCEL_PATH, temp_file)
        
        wb = openpyxl.load_workbook(temp_file, read_only=True, data_only=True)
        
        # Find all sheets starting with "Re-assign" (e.g., Re-assign, Re-assign-MD)
        reassign_sheets = [s for s in wb.sheetnames if s.startswith('Re-assign')]
        if not reassign_sheets:
            print(f"No Re-assign* sheets found in Excel. Available: {wb.sheetnames}")
            return item_status
        
        print(f"Found Re-assign sheets: {reassign_sheets}")
        
        # Read all Re-assign* sheets
        for sheet_name in reassign_sheets:
            ws = wb[sheet_name]
            current_module = None
            
            # Re-assign starts from row 3, Re-assign-MD starts from row 1
            start_row = 1 if sheet_name == 'Re-assign-MD' else 3
            
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not row or not row[0]:  # Skip empty rows
                    continue
                
                item_id = row[0]  # Column A: Item ID or Module name
                description = row[1] if len(row) > 1 else None  # Column B: Description
                owner = row[5] if len(row) > 5 else None  # Column F: Owner
                status = row[7] if len(row) > 7 else None  # Column H: Status
                
                # Convert to string if not already
                item_id_str = str(item_id).strip() if item_id else ""
                
                # Check if this row is a module header (e.g., "1.0 - LIBRARY CHECK")
                # Pattern: starts with digit, contains " - " but not an item ID prefix
                is_item_id = item_id_str.startswith(('IMP-', 'STA-', 'MIG-'))
                is_module_header = (
                    not is_item_id and 
                    item_id_str and 
                    re.match(r'^\d+\.?\d*\s*-\s*', item_id_str)  # Starts with "X.Y - "
                )
                
                if is_module_header:
                    current_module = item_id_str
                    print(f"Found module header: {current_module}")
                    continue
                
                if is_item_id:
                    # Skip if already exists (first sheet takes priority)
                    if item_id in item_status:
                        continue
                    
                    # Normalize status
                    status_normalized = 'not_started'  # Default
                    if status:
                        status_str = str(status).strip().lower()
                        if 'completed' in status_str:
                            status_normalized = 'completed'
                        elif 'in progress' in status_str or 'in-progress' in status_str:
                            status_normalized = 'in_progress'
                        elif 'not started' in status_str or 'not-started' in status_str:
                            status_normalized = 'not_started'
                    
                    # Convert module name to Check_modules directory format
                    module_dir_name = convert_excel_module_to_directory_format(current_module) if current_module else "Unknown"
                    
                    item_status[item_id] = {
                        'status': status_normalized,
                        'owner': owner if owner else 'Unassigned',
                        'module_name': module_dir_name,
                        'description': description if description else '',
                        'source': sheet_name
                    }
        
        _item_status_cache = item_status
        _item_status_cache_time = now
        _item_status_file_mtime = current_mtime  # Save file modification time
        print(f"Loaded {len(item_status)} items from Excel ({', '.join(reassign_sheets)})")
        
    except Exception as e:
        print(f"Error reading Re-assign Excel: {e}")
    finally:
        # Clean up temp file
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except:
                pass
    
    return item_status


# Cache for remote branch info (refreshed periodically)
_remote_branches_cache = {}
_cache_timestamp = None
CACHE_TTL_SECONDS = 60  # Cache for 60 seconds


async def get_remote_branches_info() -> dict:
    """Get all assignment branches info from remote using git ls-remote (fast)"""
    global _remote_branches_cache, _cache_timestamp
    
    # Check cache
    now = datetime.now()
    if _cache_timestamp and (now - _cache_timestamp).total_seconds() < CACHE_TTL_SECONDS:
        return _remote_branches_cache
    
    # Use any workspace to run git ls-remote
    workspaces = [d for d in os.listdir(WORKSPACES_DIR) 
                  if os.path.isdir(os.path.join(WORKSPACES_DIR, d))]
    if not workspaces:
        return {}
    
    workspace_path = os.path.join(WORKSPACES_DIR, workspaces[0])
    
    try:
        # git ls-remote is fast - only gets refs without downloading objects
        result = subprocess.run(
            ["git", "ls-remote", "origin"],
            cwd=workspace_path,
            capture_output=True,
            text=True,
            timeout=10
        )
        
        branches = {}
        if result.returncode == 0:
            for line in result.stdout.strip().split("\n"):
                if "refs/heads/assignment/" in line:
                    parts = line.split("\t")
                    if len(parts) == 2:
                        sha, ref = parts
                        branch_name = ref.replace("refs/heads/", "")
                        branches[branch_name] = sha
        
        _remote_branches_cache = branches
        _cache_timestamp = now
        return branches
        
    except Exception as e:
        print(f"Error getting remote branches: {e}")
        return _remote_branches_cache or {}


def get_git_info(workspace_path: str, developer_name_snake: str = "", remote_branches: dict = None) -> dict:
    """Get git information from a workspace"""
    result = {
        "current_branch": "unknown",
        "assignment_branch": None,
        "last_commit_time": None,
        "last_commit_message": None,
        "modified_files_count": 0,
        "modified_items": [],
        "remote_sha": None
    }
    
    try:
        # Get current branch
        branch_result = subprocess.run(
            ["git", "rev-parse", "--abbrev-ref", "HEAD"],
            cwd=workspace_path,
            capture_output=True,
            text=True,
            timeout=5
        )
        if branch_result.returncode == 0:
            result["current_branch"] = branch_result.stdout.strip()
        
        # Find assignment branch for this developer from remote_branches
        if remote_branches and developer_name_snake:
            assignment_branches = [
                b for b in remote_branches.keys() 
                if f"assignment/{developer_name_snake}_" in b
            ]
            if assignment_branches:
                # Sort by date (ascending) and get the first/original assignment
                assignment_branches.sort()
                result["assignment_branch"] = assignment_branches[0]
                result["remote_sha"] = remote_branches.get(assignment_branches[0])
        
        # Get last commit info - prefer assignment branch if available
        if result["assignment_branch"]:
            # Read from assignment branch (most up-to-date work)
            log_result = subprocess.run(
                ["git", "log", "-1", "--format=%ai|%s", f"origin/{result['assignment_branch']}"],
                cwd=workspace_path,
                capture_output=True,
                text=True,
                timeout=5
            )
        else:
            # Fallback to local HEAD
            log_result = subprocess.run(
                ["git", "log", "-1", "--format=%ai|%s"],
                cwd=workspace_path,
                capture_output=True,
                text=True,
                timeout=5
            )
        
        if log_result.returncode == 0 and log_result.stdout.strip():
            parts = log_result.stdout.strip().split("|", 1)
            if len(parts) >= 1:
                result["last_commit_time"] = parts[0]
            if len(parts) >= 2:
                result["last_commit_message"] = parts[1]
        
        # Get modified items from the last commit on assignment branch
        if result["assignment_branch"]:
            # Get files changed in the last commit on assignment branch
            show_result = subprocess.run(
                ["git", "show", "--name-only", "--format=", f"origin/{result['assignment_branch']}"],
                cwd=workspace_path,
                capture_output=True,
                text=True,
                timeout=5
            )
        else:
            # Fallback: get files changed in last local commit
            show_result = subprocess.run(
                ["git", "show", "--name-only", "--format=", "HEAD"],
                cwd=workspace_path,
                capture_output=True,
                text=True,
                timeout=5
            )
        
        if show_result.returncode == 0 and show_result.stdout.strip():
            files = show_result.stdout.strip().split("\n")
            # Filter only Check_modules checker files and extract unique item IDs
            item_ids = set()
            for file_path in files:
                # Only look for checker python files: checker/IMP-X-X-X-XX.py
                if ("Check_modules/" in file_path or "CHECKLIST/Check_modules/" in file_path) and \
                   "checker/" in file_path and file_path.endswith(".py"):
                    # Extract item IDs like IMP-8-0-0-01, IMP-10-0-0-00, etc.
                    match = re.search(r'(IMP-\d+-\d+-\d+-\d+|STA-\d+-\d+-\d+-\d+|MIG-\w+-\d+-\d+-\d+-\d+)', file_path)
                    if match:
                        item_ids.add(match.group(1))
            result["modified_items"] = sorted(list(item_ids))
            result["modified_files_count"] = len(item_ids)
        else:
            # Fallback: check local uncommitted changes
            status_result = subprocess.run(
                ["git", "status", "--porcelain"],
                cwd=workspace_path,
                capture_output=True,
                text=True,
                timeout=5
            )
            if status_result.returncode == 0:
                lines = [l for l in status_result.stdout.strip().split("\n") if l]
                result["modified_files_count"] = len(lines)
            
    except Exception as e:
        print(f"Error getting git info for {workspace_path}: {e}")
    
    return result


def get_recent_commits(workspace_path: str, developer_name: str, workspace_name: str, limit: int = 5) -> List[dict]:
    """Get recent commits from a workspace"""
    commits = []
    
    # Get item status from Excel
    item_status_map = get_item_status_from_excel()
    
    try:
        # Get recent commits with timestamp
        log_result = subprocess.run(
            ["git", "log", f"-{limit}", "--format=%H|%ai|%s"],
            cwd=workspace_path,
            capture_output=True,
            text=True,
            timeout=10
        )
        
        if log_result.returncode == 0 and log_result.stdout.strip():
            for line in log_result.stdout.strip().split("\n"):
                parts = line.split("|", 2)
                if len(parts) >= 3:
                    commit_hash, commit_time, commit_msg = parts
                    
                    # Parse item_id from commit message if present
                    item_id = None
                    module = None
                    status = "committed"  # Default status
                    
                    if "IMP-" in commit_msg or "STA-" in commit_msg or "MIG-" in commit_msg:
                        match = re.search(r"(IMP-\d+-\d+-\d+-\d+|STA-\d+-\d+-\d+-\d+|MIG-\w+-\d+-\d+-\d+-\d+)", commit_msg)
                        if match:
                            item_id = match.group(1)
                            # Extract module from item_id
                            parts_id = item_id.split("-")
                            if len(parts_id) >= 2:
                                module = f"{parts_id[0]}-{parts_id[1]}.0"
                            
                            # Get real status from Excel if available
                            if item_id in item_status_map:
                                status = item_status_map[item_id]['status']
                    
                    commits.append({
                        "id": commit_hash[:8],
                        "developer": developer_name,
                        "action": commit_msg,
                        "item_id": item_id,
                        "module": module,
                        "message": commit_msg,
                        "status": status,
                        "time": commit_time,
                        "workspace": workspace_name
                    })
    except Exception as e:
        print(f"Error getting commits for {workspace_path}: {e}")
    
    return commits


def is_today(time_str: str) -> bool:
    """Check if a time string is from today"""
    if not time_str:
        return False
    try:
        # Parse git date format: 2025-01-07 16:06:49 +0800
        commit_date = datetime.strptime(time_str[:10], "%Y-%m-%d").date()
        return commit_date == datetime.now().date()
    except:
        return False


@router.get("/developers", response_model=List[DeveloperStatus])
async def get_developers_status():
    """Get status of all developers from workspaces"""
    developers = []
    
    if not os.path.exists(WORKSPACES_DIR):
        return developers
    
    # Get remote branches info once (fast, using git ls-remote)
    remote_branches = await get_remote_branches_info()
    
    for workspace_name in os.listdir(WORKSPACES_DIR):
        workspace_path = os.path.join(WORKSPACES_DIR, workspace_name)
        
        if not os.path.isdir(workspace_path):
            continue
        
        # Check if it's a git repository
        git_dir = os.path.join(workspace_path, ".git")
        if not os.path.exists(git_dir):
            continue
        
        developer_name = parse_developer_name(workspace_name)
        # Extract snake_case developer name for branch matching
        match = re.match(r"(.+?)_workspace_\d+_\d+", workspace_name)
        developer_name_snake = match.group(1) if match else ""
        
        git_info = get_git_info(workspace_path, developer_name_snake, remote_branches)
        
        # Determine status based on activity
        status = "idle"
        if git_info["modified_files_count"] > 0:
            status = "working"
        elif git_info["last_commit_time"] and is_today(git_info["last_commit_time"]):
            status = "active"
        
        # Use assignment branch if available, otherwise current branch
        display_branch = git_info["assignment_branch"] or git_info["current_branch"]
        
        # Try to get assigned item from branch name
        # Branch format: assignment/zhongyu_sun_20260106
        assigned_item = None
        if display_branch and display_branch not in ["master", "main", "unknown"]:
            # Extract assignment info from branch name
            if display_branch.startswith("assignment/"):
                assigned_item = display_branch.replace("assignment/", "")
            else:
                assigned_item = display_branch
        
        developers.append(DeveloperStatus(
            developer_name=developer_name,
            workspace_name=workspace_name,
            current_branch=display_branch,
            last_commit_time=git_info["last_commit_time"],
            last_commit_message=git_info["last_commit_message"],
            modified_files_count=git_info["modified_files_count"],
            modified_items=git_info.get("modified_items", []),
            status=status,
            assigned_item=assigned_item
        ))
    
    # Sort by status (working > active > idle), then by name
    status_order = {"working": 0, "active": 1, "idle": 2}
    developers.sort(key=lambda d: (status_order.get(d.status, 3), d.developer_name))
    
    return developers


@router.get("/activity", response_model=List[ActivityItem])
async def get_recent_activity(limit: int = 20):
    """Get recent activity across all workspaces"""
    all_activities = []
    
    if not os.path.exists(WORKSPACES_DIR):
        return all_activities
    
    for workspace_name in os.listdir(WORKSPACES_DIR):
        workspace_path = os.path.join(WORKSPACES_DIR, workspace_name)
        
        if not os.path.isdir(workspace_path):
            continue
        
        git_dir = os.path.join(workspace_path, ".git")
        if not os.path.exists(git_dir):
            continue
        
        developer_name = parse_developer_name(workspace_name)
        commits = get_recent_commits(workspace_path, developer_name, workspace_name, limit=5)
        
        for commit in commits:
            all_activities.append(ActivityItem(
                id=commit["id"],
                developer=commit["developer"],
                action=commit.get("message", "commit"),
                item_id=commit.get("item_id"),
                module=commit.get("module"),
                status=commit["status"],
                time=commit["time"],
                workspace=commit["workspace"]
            ))
    
    # Sort by time (most recent first)
    all_activities.sort(key=lambda a: a.time, reverse=True)
    
    return all_activities[:limit]


@router.get("/items")
async def get_all_items():
    """Get all items from Re-assign Excel with their status"""
    items = []
    
    # Get item status from Excel
    item_status_map = get_item_status_from_excel()
    
    # Convert to list format
    for item_id, info in item_status_map.items():
        items.append({
            "item_id": item_id,
            "owner": info.get("owner"),
            "status": info.get("status"),
            "module_name": info.get("module_name"),
            "description": info.get("description"),
            "source": info.get("source", "Re-assign")  # PD or MD
        })
    
    return items


@router.post("/items/refresh")
async def refresh_items():
    """Clear cache and reload items from Excel"""
    clear_item_status_cache()
    return {"message": "Cache cleared", "items_count": len(get_item_status_from_excel())}


@router.get("/stats", response_model=DashboardStats)
async def get_dashboard_stats():
    """Get dashboard statistics"""
    total_developers = 0
    active_developers = 0
    completed_today = 0
    total_commits_today = 0
    
    if not os.path.exists(WORKSPACES_DIR):
        return DashboardStats(
            total_developers=0,
            active_developers=0,
            completed_today=0,
            total_commits_today=0
        )
    
    for workspace_name in os.listdir(WORKSPACES_DIR):
        workspace_path = os.path.join(WORKSPACES_DIR, workspace_name)
        
        if not os.path.isdir(workspace_path):
            continue
        
        git_dir = os.path.join(workspace_path, ".git")
        if not os.path.exists(git_dir):
            continue
        
        total_developers += 1
        
        # Check today's commits
        try:
            today = datetime.now().strftime("%Y-%m-%d")
            log_result = subprocess.run(
                ["git", "log", f"--since={today} 00:00", "--oneline"],
                cwd=workspace_path,
                capture_output=True,
                text=True,
                timeout=5
            )
            
            if log_result.returncode == 0 and log_result.stdout.strip():
                commits = [l for l in log_result.stdout.strip().split("\n") if l]
                commit_count = len(commits)
                if commit_count > 0:
                    active_developers += 1
                    total_commits_today += commit_count
                    completed_today += commit_count
        except Exception as e:
            print(f"Error getting stats for {workspace_path}: {e}")
    
    return DashboardStats(
        total_developers=total_developers,
        active_developers=active_developers,
        completed_today=completed_today,
        total_commits_today=total_commits_today
    )


@router.post("/refresh")
async def refresh_from_remote():
    """Fetch latest from remote for all workspaces"""
    global _remote_branches_cache, _cache_timestamp
    
    if not os.path.exists(WORKSPACES_DIR):
        return {"status": "error", "message": "Workspaces directory not found"}
    
    workspaces = [d for d in os.listdir(WORKSPACES_DIR) 
                  if os.path.isdir(os.path.join(WORKSPACES_DIR, d))]
    
    if not workspaces:
        return {"status": "error", "message": "No workspaces found"}
    
    # Fetch all workspaces in parallel (up to 5 at a time)
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    
    def fetch_workspace(workspace_name):
        workspace_path = os.path.join(WORKSPACES_DIR, workspace_name)
        git_dir = os.path.join(workspace_path, ".git")
        if not os.path.exists(git_dir):
            return {"workspace": workspace_name, "status": "skip", "reason": "not a git repo"}
        
        try:
            result = subprocess.run(
                ["git", "fetch", "origin"],
                cwd=workspace_path,
                capture_output=True,
                text=True,
                timeout=30
            )
            return {
                "workspace": workspace_name,
                "status": "success" if result.returncode == 0 else "error",
                "output": result.stderr if result.returncode != 0 else "OK"
            }
        except subprocess.TimeoutExpired:
            return {"workspace": workspace_name, "status": "timeout"}
        except Exception as e:
            return {"workspace": workspace_name, "status": "error", "error": str(e)}
    
    # Fetch workspaces with ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=5) as executor:
        results = list(executor.map(fetch_workspace, workspaces))
    
    # Clear cache to force refresh
    _cache_timestamp = None
    _remote_branches_cache = {}
    
    success_count = sum(1 for r in results if r["status"] == "success")
    
    return {
        "status": "success",
        "message": f"Fetched {success_count}/{len(workspaces)} workspaces successfully",
        "results": results
    }


@router.get("/excel-summary")
async def get_excel_summary():
    """Get summary statistics from Excel file (ME_PD_1 and Re-assign tabs)"""
    if not os.path.exists(REASSIGN_EXCEL_PATH):
        raise HTTPException(status_code=404, detail="Excel file not found")
    
    import shutil
    import tempfile
    import uuid
    
    temp_file = None
    try:
        # Copy to temp file to avoid lock issues
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"excel_summary_{uuid.uuid4().hex}.xlsx")
        shutil.copy2(REASSIGN_EXCEL_PATH, temp_file)
        
        wb = openpyxl.load_workbook(temp_file, read_only=True, data_only=True)
        
        # 1. Read ME_PD_1 TAB for total modules and items
        total_modules = 0
        total_items = 0
        
        if 'ME_PD_1' in wb.sheetnames:
            ws = wb['ME_PD_1']
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                
                item_id_str = str(row[0]).strip()
                
                # Check if module header (e.g., " 1.0 - LIBRARY CHECK")
                is_item_id = item_id_str.startswith(('IMP-', 'STA-', 'MIG-'))
                is_module_header = (
                    not is_item_id and 
                    item_id_str and 
                    re.match(r'^\s*\d+\.?\d*\s*-\s*', item_id_str)
                )
                
                if is_module_header:
                    total_modules += 1
                elif is_item_id:
                    total_items += 1
        
        # 2. Read Re-assign* TABs for developer statistics
        reassign_sheets = [s for s in wb.sheetnames if s.startswith('Re-assign')]
        developer_stats = {}
        status_distribution = {
            'completed': 0,
            'in_progress': 0,
            'not_started': 0
        }
        
        for sheet_name in reassign_sheets:
            ws = wb[sheet_name]
            start_row = 1 if sheet_name == 'Re-assign-MD' else 3
            
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not row or not row[0]:
                    continue
                
                item_id = str(row[0]).strip() if row[0] else ""
                owner = str(row[5]).strip() if len(row) > 5 and row[5] else None
                status = str(row[7]).strip().lower() if len(row) > 7 and row[7] else ""
                
                # Only process item rows
                if not item_id.startswith(('IMP-', 'STA-', 'MIG-')):
                    continue
                
                # Normalize status
                status_normalized = 'not_started'
                if 'completed' in status:
                    status_normalized = 'completed'
                elif 'in progress' in status or 'in-progress' in status:
                    status_normalized = 'in_progress'
                
                # Update status distribution
                status_distribution[status_normalized] += 1
                
                # Update developer statistics
                if owner and owner != 'Unassigned':
                    if owner not in developer_stats:
                        developer_stats[owner] = {
                            'total_assigned': 0,
                            'completed': 0,
                            'in_progress': 0,
                            'not_started': 0
                        }
                    
                    developer_stats[owner]['total_assigned'] += 1
                    developer_stats[owner][status_normalized] += 1
        
        wb.close()
        
        # Convert developer_stats to list format
        developer_list = []
        for dev_name, stats in developer_stats.items():
            completion_rate = (stats['completed'] / stats['total_assigned'] * 100) if stats['total_assigned'] > 0 else 0
            developer_list.append({
                'name': dev_name,
                'total_assigned': stats['total_assigned'],
                'completed': stats['completed'],
                'in_progress': stats['in_progress'],
                'not_started': stats['not_started'],
                'completion_rate': round(completion_rate, 1)
            })
        
        # Sort by total_assigned descending
        developer_list.sort(key=lambda x: x['total_assigned'], reverse=True)
        
        return {
            'total_modules': total_modules,
            'total_items': total_items,
            'status_distribution': status_distribution,
            'developers': developer_list,
            'excel_path': REASSIGN_EXCEL_PATH
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading Excel: {str(e)}")
    finally:
        # Clean up temp file
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except:
                pass
