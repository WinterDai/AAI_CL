import openpyxl
import sys

excel_path = r"C:\Users\yuyin\OneDrive - Cadence Design Systems Inc\2025\Checklist\RELEASE\DOC\DR3_SSCET_BE_Check_List_v0.2_20260105.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['Re-assign']
    
    print("Columns:")
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        print(row)
    
    print("\nFirst 15 rows:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=16, values_only=True)):
        print(f"{i+2}: {row}")
    
    print(f"\nTotal rows: {ws.max_row}")
    
except Exception as e:
    print(f"Error: {e}")
    sys.exit(1)
