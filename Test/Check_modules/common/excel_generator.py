#!/usr/bin/env python3
################################################################################
# Script Name: excel_generator.py
#
# Purpose:
#   Convert per-module summary YAML into formatted Excel workbook.
#   Applies styling, auto-sizing, and writes to Work/Results/<module>/<module>.xlsx.
#
# Usage:
#   python excel_generator.py -summary <SUMMARY_YAML> -root <ROOT>
#   (Normally invoked indirectly via check_flowtool.py)
#
# Author: yyin
# Date:   2025-10-23
################################################################################
"""Generate an Excel (.xlsx) summary from a module summary YAML.

Output Path Pattern:
    <work_root>/Work/Results/<module>/<module>.xlsx

Notes:
    - This script now ALWAYS writes .xlsx (previous CSV fallback removed).
    - Requires 'openpyxl'. Install if missing:
            pip install openpyxl
    - Column widths auto-sized, header row bold + centered, auto-filter enabled.
"""
from pathlib import Path
from typing import Any, Dict, List, Optional
import argparse
import yaml
import sys


def load_summary(summary_yaml: Path) -> Dict[str, Any]:
    if not summary_yaml.is_file():
        raise FileNotFoundError(f"Summary YAML not found: {summary_yaml}")
    with summary_yaml.open("r") as f:
        return yaml.safe_load(f) or {}


def ensure_results_dir(work_root: Path, module: str) -> Path:
    """
    Create directory: <work_root>/Work/Results/<module>
    """
    out_dir = work_root / "Work" / "Results" / module
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _load_openpyxl():
    """Import and return needed openpyxl symbols.

    Raises ImportError with helpful message if unavailable.
    """
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
        from openpyxl.styles import Font, Alignment, Border, Side  # type: ignore
        return Workbook, get_column_letter, Font, Alignment, Border, Side
    except Exception as e:  # noqa: BLE001 broad for user friendliness
        raise ImportError(
            "openpyxl is required to generate .xlsx files. Install via 'pip install openpyxl'. Original error: "
            + str(e)
        )


def auto_size_columns(ws):
    # Adjusted to work only if openpyxl loaded
    try:
        from openpyxl.utils import get_column_letter  # type: ignore
    except (ImportError, SyntaxError):
        return
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def has_warnings(item_info: Dict[str, Any]) -> bool:
    """Check if item has any warnings (non-empty warnings list excluding occurrence-only dict)."""
    warnings = item_info.get('warnings') or []
    for w in warnings:
        if isinstance(w, dict) and not (len(w) == 1 and 'occurrence' in w):
            return True
    return False


def build_rows(summary: Dict[str, Any]) -> List[List[Any]]:
    """Construct row data including failures and warnings.

    New column order:
        Module,Stage,ItemID,Executed,Status,Description,Kind,Occurrence,Index,Detail,SourceLine,SourceFile,Reason

    Kind: FAILURE or WARNING (blank when there are no detail rows)
    Occurrence: Uses the occurrence value for the corresponding kind (failures or warnings).
    Index: The index inside the failures/warnings list entries.
    """
    module = summary.get("module", "")
    stage = summary.get("stage", "")
    check_items: Dict[str, Any] = summary.get("check_items", {})
    rows: List[List[Any]] = []
    for item_id, item_data in check_items.items():
        executed = item_data.get("executed")
        # Keep original status (don't override with WARN)
        status = str(item_data.get("status", "")).upper()
        desc = item_data.get("description", "")

        # Process failures
        failures = item_data.get("failures", []) or []
        fail_occurrence = ""
        norm_failures: List[Dict[str, Any]] = []
        for f in failures:
            if isinstance(f, dict) and "occurrence" in f and len(f.keys()) == 1:
                fail_occurrence = f.get("occurrence", "")
            else:
                norm_failures.append(f)
        if norm_failures:
            for failure in norm_failures:
                rows.append([
                    module, stage, item_id, executed, status, desc, "FAILURE", fail_occurrence,
                    failure.get("index", ""),
                    failure.get("detail", ""),
                    failure.get("source_line", ""),
                    failure.get("source_file", ""),
                    failure.get("reason", "")
                ])

        # Process warnings
        warnings = item_data.get("warnings", []) or []
        warn_occurrence = ""
        norm_warnings: List[Dict[str, Any]] = []
        for w in warnings:
            if isinstance(w, dict) and "occurrence" in w and len(w.keys()) == 1:
                warn_occurrence = w.get("occurrence", "")
            else:
                norm_warnings.append(w)
        if norm_warnings:
            for warning in norm_warnings:
                rows.append([
                    module, stage, item_id, executed, status, desc, "WARNING", warn_occurrence,
                    warning.get("index", ""),
                    warning.get("detail", ""),
                    warning.get("source_line", ""),
                    warning.get("source_file", ""),
                    warning.get("reason", "")
                ])

        # If neither failures nor warnings produced any detail rows, add a single blank row for the item
        if not norm_failures and not norm_warnings:
            rows.append([module, stage, item_id, executed, status, desc, "", "", "", "", "", "", ""])
    return rows


def write_excel(summary: Dict[str, Any], out_dir: Path) -> Optional[Path]:
    Workbook, get_column_letter, Font, Alignment, Border, Side = _load_openpyxl()

    module = summary.get("module", "UNKNOWN_MODULE")
    xlsx_path = out_dir / f"{module}.xlsx"
    headers = [
        "Module","Stage","ItemID","Executed","Status","Description","Kind","Occurrence",
        "Index","Detail","SourceLine","SourceFile","Reason"
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
        c.alignment = center
        c.border = thin_border
    data_rows = build_rows(summary)
    for row in data_rows:
        ws.append(row)

    # Conditional formatting + row highlight for FAIL
    from openpyxl.styles import PatternFill  # type: ignore
    from openpyxl.formatting.rule import CellIsRule  # type: ignore
    
    status_col_letter = get_column_letter(headers.index("Status") + 1)
    status_range = f"{status_col_letter}2:{status_col_letter}{ws.max_row}"
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ws.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"PASS"'], fill=pass_fill))
    ws.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"FAIL"'], fill=fail_fill))
    ws.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"WARN"'], fill=warn_fill))
    # Row highlight for PASS and FAIL rows and apply borders
    status_col_idx = headers.index("Status") + 1
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=status_col_idx).value
        if cell_val:
            cell_val_upper = str(cell_val).upper()
            if cell_val_upper.startswith("FAIL"):
                for c in ws[r]:
                    c.fill = fail_fill
                    c.border = thin_border
            elif cell_val_upper == "PASS":
                for c in ws[r]:
                    c.fill = pass_fill
                    c.border = thin_border
            elif cell_val_upper == "WARN":
                for c in ws[r]:
                    c.fill = warn_fill
                    c.border = thin_border
            else:
                for c in ws[r]:
                    c.border = thin_border
        else:
            for c in ws[r]:
                c.border = thin_border

    # Overview sheet (unique items statistics, not expanded by failures)
    overview = wb.create_sheet(title="Overview")
    overview.append(["Metric", "Count"])
    overview["A1"].font = bold; overview["B1"].font = bold
    overview["A1"].border = thin_border; overview["B1"].border = thin_border
    check_items: Dict[str, Any] = summary.get("check_items", {})
    total_items = len(check_items)
    executed_items = 0
    pass_items = 0
    fail_items = 0
    warn_items = 0
    total_failures = 0  # aggregate all failure entries across items (exclude occurrence record)
    total_warnings = 0  # aggregate all warning entries across items (exclude occurrence record)
    for item_id, item_data in check_items.items():
        executed_flag = str(item_data.get("executed", "")).upper() in {"YES", "TRUE", "1"}
        if executed_flag:
            executed_items += 1
        
        # Count based on actual status
        status_val = str(item_data.get("status", "")).upper()
        if status_val.startswith("FAIL"):
            fail_items += 1
        elif status_val == "PASS":
            pass_items += 1
        
        # ADDITIONAL: if item has warnings, also count as warn (doesn't override status)
        if has_warnings(item_data):
            warn_items += 1
        # Count only detailed entries (exclude occurrence-only dict)
        failures_list = item_data.get("failures", []) or []
        warnings_list = item_data.get("warnings", []) or []
        total_failures += sum(1 for f in failures_list if isinstance(f, dict) and not (len(f) == 1 and 'occurrence' in f))
        total_warnings += sum(1 for w in warnings_list if isinstance(w, dict) and not (len(w) == 1 and 'occurrence' in w))
    overview_rows = [
        ["Total items", total_items],
        ["Executed", executed_items],
        ["PASS", pass_items],
        ["FAIL", fail_items],
        ["WARN", warn_items],
        ["Total failure entries", total_failures],
        ["Total warning entries", total_warnings],
    ]
    for r in overview_rows:
        overview.append(r)
    # Apply borders to all data rows in overview
    for row in overview.iter_rows(min_row=2, max_row=overview.max_row):
        for cell in row:
            cell.border = thin_border
    # Autosize overview columns
    for col in overview.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            l = len(str(cell.value))
            if l > max_len:
                max_len = l
        overview.column_dimensions[letter].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_size_columns(ws)
    wb.save(xlsx_path)
    print(f"[INFO] Excel written: {xlsx_path}")
    return xlsx_path


# CSV fallback removed intentionally. If needed in future, restore previous implementation.


def summary_yaml_to_excel(summary_yaml: Path, work_root: Path) -> Optional[Path]:
    summary = load_summary(summary_yaml)
    module = summary.get("module", "UNKNOWN_MODULE")
    out_dir = ensure_results_dir(work_root, module)
    return write_excel(summary, out_dir)



def _cli():
    parser = argparse.ArgumentParser(description="Generate <Work>/Results/<module>/<module>.xlsx from summary YAML.")
    parser.add_argument("-result_yaml", help="Path to module summary YAML (e.g. Check_modules/<mod>/outputs/<mod>.yaml)")
    parser.add_argument("-work_root", "-root", default="..", help="Root directory of checklist project (default ..)")
    args = parser.parse_args()
    summary_path = Path(args.result_yaml).resolve()
    work_root = Path(args.work_root).resolve()
    if not summary_path.is_file():
        print(f"[ERROR] Summary YAML not found: {summary_path}")
        return 2
    try:
        produced = summary_yaml_to_excel(summary_path, work_root)
        if produced:
            print(f"[INFO] XLSX generated: {produced}")
        return 0
    except Exception as e:
        print(f"[ERROR] Failed: {e}")
        return 3


if __name__ == "__main__":  # pragma: no cover
    sys.exit(_cli())

